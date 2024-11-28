import os
from openpyxl import load_workbook, Workbook
from app import db
from app.models import Local, Peca  # Ajuste o import conforme a estrutura da sua aplicação
from app.management.lista_lideres import lideres, get_lider

# Criação do diretório de saída
output_dir = "listas_geradas"
os.makedirs(output_dir, exist_ok=True)

# Caminho para o arquivo modelo
template_path = "excel_dados_inventario/Lista - Inventário 2024.xlsx"

def generate_inventory_lists():
    locais = Local.query.all()

    for local in locais:
        piece_count = 0  # Contador de peças
        file_index = 1   # Índice do arquivo
        wb = None
        ws = None

        for peca in local.pecas:
            # Se contador é 0 ou múltiplo de 20, cria novo arquivo
            if piece_count % 20 == 0:
                if wb:
                    # Salva o arquivo anterior
                    output_file = os.path.join(output_dir, f"{local.nome}_{file_index}_{local.almoxarifado}.xlsx")
                    wb.save(output_file)
                    print(f"Arquivo gerado: {output_file}")
                    file_index += 1  # Incrementa o índice do arquivo

                # Cria um novo workbook e worksheet
                wb = load_workbook(template_path)
                ws = wb.active

                # Preenche o cabeçalho para o novo arquivo
                lider = get_lider(local.almoxarifado, local.nome)
                ws["B2"] = lider  # Nome do líder
                ws["B4"] = local.almoxarifado  # Almoxarifado
                ws.merge_cells("B4:D4")  # Mesclar B4, C4, e D4
                ws["F4"] = local.nome  # Nome do Local

                start_row = 7  # Reinicia a escrita de peças na linha 7

            # Adiciona a peça na planilha
            row = start_row + (piece_count % 20)  # Calcula a linha com base no contador
            ws[f"A{row}"] = peca.codigo  # Código da peça
            ws[f"B{row}"] = peca.descricao  # Descrição da peça
            ws.merge_cells(f"B{row}:E{row}")  # Mesclar B7, C7, D7, e E7
            ws[f"G{row}"] = local.estante  # Estante

            piece_count += 1  # Incrementa o contador de peças

        # Salva o último arquivo restante
        if wb:
            output_file = os.path.join(output_dir, f"{local.nome}_{file_index}_{local.almoxarifado}.xlsx")
            wb.save(output_file)
            print(f"Arquivo gerado: {output_file}")